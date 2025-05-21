# solution_utils.py
from datetime import datetime

import pandas as pd
from gurobipy import Model, GRB
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from Code.input_module import load_all_inputs


def parse_solution(model: Model, vars: dict, r_p_k: dict, P: list, K: list) -> dict:

    """
    Parseer een opgeloste Gurobi-model naar een dictionaryvorm:
    {"x": {(w, p, s): 1.0}, ...} met alleen niet-nulwaarden.

    Parameters:
    - model: opgeloste Gurobi model
    - vars: dict met Gurobi-variabelen (zoals uit build_gurobi_model)

    Returns:
    - dict[str, dict[tuple, float]]: parsed oplossing
    """
    if model.status not in [GRB.OPTIMAL, GRB.TIME_LIMIT] or model.SolCount == 0:
        print("⚠️ Geen bruikbare oplossing beschikbaar.")
        return {}

    parsed = {}


    for name, varset in vars.items():
        entries = {
            idx: var.X
            for idx, var in varset.items()
            if isinstance(var.X, float) and abs(var.X) > 1e-6
        }
        parsed[name] = entries

    # Extra post-processing: q[w, p, s] ≥ q[w, p, s-1]
    if "q" in parsed:
        q_vals = parsed["q"]
        # verzamel alle combinaties (w, p) en sorteer hun shifts
        from collections import defaultdict
        by_worker_pos = defaultdict(list)
        for (w, p, s) in q_vals:
            by_worker_pos[(w, p)].append(s)

        # voor elke (w, p), propagate q=1 forward
        for (w, p), shifts in by_worker_pos.items():
            shifts_sorted = sorted(shifts)
            q_previous = 0
            for s in shifts_sorted:
                key = (w, p, s)
                val = q_vals.get(key, 0)
                if val >= 0.5 or q_previous >= 0.5:
                    q_vals[key] = 1.0
                    q_previous = 1.0
        #  Zelf q berekenen op basis van h
        h = parsed.get("h", {})
        if h:  # Alleen herberekenen als 'h' beschikbaar is
            q_manual = {}
            K_p = {p: [k for (pp, k) in r_p_k if pp == p] for p in P}
            for (w, p, s) in vars["x"]:
                required_ks = K_p.get(p, [])
                if not required_ks:
                    continue
                qualified = all(h.get((w, k, s), 0) >= 0.99 for k in required_ks)
                q_manual[(w, p, s)] = 1.0 if qualified else 0.0
            parsed["q"] = q_manual
        return parsed


def parse_solution_with_tvc(model, vars, inputs, r_p_k: dict, P: list, K: list):
    parsed = parse_solution(model, vars, r_p_k, P, K)
    if "q" in parsed and "x" in parsed:
        parsed["tvc"] = calculate_team_flexibility_parsed(
            parsed_solution=parsed,
            Pm=inputs["Pm"],
            P_map=inputs["P_map"],
            q=parsed["q"],
            f_w_s=inputs["f_w_s"]
        )

        return parsed

    print("Parsed solution keys:", parsed.keys())
    return parsed

start_time_global = datetime.now()
def time_since_start():
    return str(datetime.now() - start_time_global).split(".")[0]  # zonder microseconden

def export_solution_to_excel(parsed_solution: dict, output_path: str) -> None:
    """
    Exporteer een parsed oplossing (zoals uit parse_solution) naar Excel.

    Parameters:
    - parsed_solution: dict met keys als "x", "y", ... en values als {(i,j,k): val}
    - output_path: pad naar outputbestand (bv. "oplossing.xlsx")
    """
    with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
        for name, values in parsed_solution.items():
            if not values:
                continue
            df = pd.DataFrame([
                (*key, value) for key, value in values.items()
            ])
            df.columns = [f"Index{i+1}" for i in range(df.shape[1] - 1)] + ["Value"]
            df.to_excel(writer, sheet_name=name, index=False)
    print(f"✅ Oplossing geëxporteerd naar: {output_path}")

def extract_position_qualification_timeline(parsed_solution: dict, P_map: dict, K_p: dict) -> dict[int, list[tuple[int, str, int]]]:
    """
    Bepaalt voor elke werknemer in welke volgorde hij een nieuwe kwalificatie behaalde,
    gebaseerd op h-status (skills), onafhankelijk van q die mogelijks achterloopt.

    Parameters:
    - parsed_solution: dictionary met parsed oplossing (moet "q" en "h" bevatten)
    - P_map: mapping van positie-id naar naam
    - K_p: mapping van posities naar vereiste skill-id's

    Returns:
    - dict[werknemer] → list van tuples (positie_id, positie_naam, shift)
    """
    q = parsed_solution.get("q", {})
    h = parsed_solution.get("h", {})

    timeline = {}

    # Stap 1: bepaal welke (w, p) initieel volledig gekwalificeerd waren via h
    initial_qualified = set()
    for (w, p, s), _ in q.items():
        if s != 0:
            continue
        required_skills = K_p.get(p, [])
        if all(h.get((w, k, 0), 0) > 0.5 for k in required_skills):
            initial_qualified.add((w, p))

    # Stap 2: filter q op nieuwe kwalificaties (niet al aanwezig in s=0)
    for (w, p, s), val in q.items():
        if val <= 0.5 or (w, p) in initial_qualified:
            continue

        if w not in timeline:
            timeline[w] = {}
        if p not in timeline[w] or s < timeline[w][p]:
            timeline[w][p] = s

    # Stap 3: sorteer per werknemer
    timeline_sorted = {}
    for w, positions in timeline.items():
        sorted_entries = sorted(positions.items(), key=lambda x: x[1])
        timeline_sorted[w] = [(p, P_map[p], s) for p, s in sorted_entries]

    return timeline_sorted

def export_qualification_timeline(timeline: dict, output_path: str):
    """
    Exporteert de kwalificatievolgorde van elke werknemer naar Excel.

    Parameters:
    - timeline: dict van werknemer naar lijst van (positie_id, positie_naam, shift)
    - output_path: pad naar Excelbestand
    """
    rows = []
    for w, entries in timeline.items():
        for i, (p, name, s) in enumerate(entries, start=1):
            rows.append({
                "Employee": w,
                "Step": i,
                "PositionID": p,
                "PositionName": name,
                "AcquiredAtShift": s
            })

    df = pd.DataFrame(rows)
    df.to_excel(output_path, index=False)
    print(f"✅ Kwalificatievolgorde geëxporteerd naar: {output_path}")

def save_solution_to_sol_file(parsed_solution: dict, output_path: str) -> None:
    """
    Schrijf een parsed oplossing weg naar een .sol-bestand zoals Gurobi dat verwacht.

    Parameters:
    - parsed_solution: dictionary uit parse_solution
    - output_path: string naar .sol-bestand
    """
    with open(output_path, "w") as f:
        f.write("# Parsed Gurobi Solution\n")
        for varname, entries in parsed_solution.items():
            for index, value in entries.items():
                index_str = ",".join(str(i) for i in index)
                f.write(f"{varname}[{index_str}] {value:.4f}\n")
    print(f"📄 .sol-bestand opgeslagen als: {output_path}")
def build_schedule_matrix(
    var_data: dict[tuple, float],
    axis: str = "shift_by_worker",
    label_fn=None
) -> pd.DataFrame:
    """
    Converteer een parsed variabele (zoals 'x') naar een rooster-DataFrame.

    Parameters:
    - var_data: dict van (w, p, s) → value (zoals parsed_solution["x"])
    - axis: "shift_by_worker" of "worker_by_shift"
    - label_fn: functie om van (w, p, s) een label te maken (bv. P_map[p]) — standaard: geef p

    Returns:
    - pd.DataFrame met bijv. shifts als rijen, werknemers als kolommen
    """
    from collections import defaultdict

    matrix = defaultdict(dict)

    for key, val in var_data.items():
        if len(key) == 3:
            w, p, s = key
        elif len(key) == 2:
            w, s = key
            p = None
        else:
            continue

        if val < 0.5:
            continue

        label = label_fn(w, p, s) if label_fn else p

        if axis == "shift_by_worker":
            matrix[s][w] = label
        else:
            matrix[w][s] = label

    df = pd.DataFrame(matrix).sort_index()
    return df
def extract_underqualified_counts(parsed_solution, P_map, sickdays: dict[int, set[int]]) -> pd.DataFrame:
    """
    Tel het aantal underqualified assignments per positie, rekening houdend met ziektedagen.
    """
    assignments = parsed_solution["x"]
    qualifications = parsed_solution.get("q", {})  # Gurobi variable "q"

    counts = {}

    for (w, p, s), assigned in assignments.items():
        if assigned < 0.5:
            continue  # niet effectief ingepland

        if s in sickdays.get(w, set()):
            continue  # werknemer was ziek → telt niet mee

        q_val = qualifications.get((w, p, s), 0)
        underqualified = assigned * (1 - q_val)  # zou 1 zijn als wel toegewezen, maar niet gekwalificeerd

        if underqualified >= 0.5:
            pos_name = P_map.get(p, f"Pos {p}")
            counts[pos_name] = counts.get(pos_name, 0) + 1

    return pd.DataFrame([{"Position": pos, "UnqualifiedCount": int(count)} for pos, count in counts.items()])


def parse_multiple_solutions(model, vars, r_p_k, P, K):

    num_solutions = model.SolCount
    solutions = []

    for i in range(num_solutions):
        model.setParam(GRB.Param.SolutionNumber, i)
        parsed = parse_solution(model, vars, r_p_k, P, K)
        solutions.append(parsed)

    return solutions



from simulator_module import build_assignment_calendar, apply_sickdays_and_replacements

def generate_sickday_calendar_detailed(parsed_solution: dict, sickdays: dict[int, set[int]], replacements: pd.DataFrame) -> pd.DataFrame:
    """
    Genereert een gedetailleerde kalender na het toepassen van ziektedagen en vervangingen,
    gebruikmakend van de uniforme STATUS codes en bestaande simulator_module functies.

    Parameters:
    - parsed_solution: parsed gurobi oplossing {"x": ..., "y": ..., "z": ...}
    - sickdays: dict {werknemer_id: set van zieke shifts}
    - replacements: DataFrame met ["Shift", "Replaced", "Replacement", "Position", "Type"]

    Returns:
    - pd.DataFrame: gedetailleerde kalender (shift × werknemer)
    """
    # Stap 1: Bouw initiële assignment kalender
    calendar = build_assignment_calendar(parsed_solution)

    # Stap 2: Pas ziekte en vervangingen toe
    updated_calendar = apply_sickdays_and_replacements(calendar, sickdays, replacements, parsed_solution)

    return updated_calendar



def export_styled_calendar_detailed(calendar_df, parsed_solution, file_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sickday Calendar"

    colors = {
        "Y": "FFFF00",
        "G": "90EE90",
        "R": "FFA500",
        "X": "B0B0B0",
        "O": "FFFFFF",
        "r_reserve": "FF0000",
        "r_training": "FFFF00",
        "r": "FFFF00"
    }

    ws.append(["Shift"] + list(calendar_df.columns))

    for row_idx, shift in enumerate(calendar_df.index, start=2):
        ws.cell(row=row_idx, column=1, value=shift)
        for col_idx, worker in enumerate(calendar_df.columns, start=2):
            val = calendar_df.at[shift, worker]
            if not isinstance(val, tuple):
                val = ("", None)

            code, data = val
            display = ""
            fill_color = None
            font = Font()

            if code == "Y":
                display = str(data)
                fill_color = "FFFF00"
            elif code == "G":
                display = str(data)
                fill_color = "90EE90"
            elif code == "R":
                display = ""
                fill_color = "FFA500"
            elif code == "X":
                display = str(data)
                fill_color = "B0B0B0"
                font = Font(bold=True)
            elif code == "O":
                display = str(data)
                fill_color = "FFFFFF"
            elif code == "r_reserve":
                display = str(data)
                fill_color = "FF0000"
                font = Font(italic=True)
            elif code == "r_training":
                pos, from_train = data
                display = f"{pos} ({from_train})"
                fill_color = "FFFF00"
            elif code == "r":
                display = str(data)
                fill_color = "FFFF00"

            if data is not None and isinstance(data, int):
                if (worker, data, shift) not in parsed_solution["q"]:
                    font.underline = "single"

            cell = ws.cell(row=row_idx, column=col_idx, value=display)
            if fill_color:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.font = font
            cell.alignment = Alignment(horizontal="center")

    # Legende
    ws_legend = wb.create_sheet("Legend")
    legend = [
        ("Code", "Meaning"),
        ("Yellow with number", "Working (position number)"),
        ("Green with number", "Training for position"),
        ("Orange", "Reserve"),
        ("Grey with bold", "Sick (original position in bold)"),
        ("White with number", "Originally assigned but removed (sick)"),
        ("Red italic", "Replacement from reserve"),
        ("Yellow 'X (Y)'", "Replacement from training — X=new, Y=original training"),
        ("Underlined", "Not qualified for that position")
    ]
    for row in legend:
        ws_legend.append(row)

    wb.save(file_path)
    return file_path
def export_sickness_logs(replacements_df, sickdays_dict, missed_trainings_dict, updated_fws_dict, export_path_prefix):
    with pd.ExcelWriter(f"{export_path_prefix}_sickness_log.xlsx", engine="xlsxwriter") as writer:
        replacements_df.to_excel(writer, sheet_name="Replacements", index=False)

        sickdays_df = pd.DataFrame([
            {"Employee": w, "SickShift": s}
            for w, shifts in sickdays_dict.items()
            for s in sorted(shifts)
        ])
        sickdays_df.to_excel(writer, sheet_name="Sickdays", index=False)

        missed_df = pd.DataFrame([
            {"Employee": w, "TrainingPosition": p, "OriginalShift": s}
            for w, lst in missed_trainings_dict.items()
            for (p, s) in sorted(lst)
        ])
        missed_df.to_excel(writer, sheet_name="MissedTrainings", index=False)

        fws_df = pd.DataFrame([
            {"Employee": w, "RecoveredLeaveShift": s}
            for w, shifts in updated_fws_dict.items()
            for s, val in shifts.items() if val > 0
        ])
        fws_df.to_excel(writer, sheet_name="LeaveRecovery", index=False)
def analyze_training_recovery(missed_trainings: dict, updated_training: dict) -> pd.DataFrame:
    """
    Controleert of gemiste trainingen zijn ingehaald, en zo ja op welk moment.
    Parameters:
    - missed_trainings: {w: set((p, s))} met gemiste (positie, shift)
    - updated_training: dict[(w, p, s)] → 1.0 voor ingeplande trainingen

    Returns:
    - pd.DataFrame met kolommen: Werknemer, Positie, GemistOp, IngehaaldOp, Status
    """
    results = []
    for w, missings in missed_trainings.items():
        for (p, s_missed) in missings:
            recovery_shift = next(
                (s_new for (ww, pp, s_new), val in updated_training.items()
                 if ww == w and pp == p and val > 0.5 and s_new > s_missed),
                None
            )
            results.append({
                "Werknemer": w,
                "Positie": p,
                "GemistOp": s_missed,
                "IngehaaldOp": recovery_shift,
                "Status": "✅ Ingehaald" if recovery_shift is not None else "❌ Niet ingehaald"
            })
    return pd.DataFrame(results)
def debug_initial_qualifications(parsed_solution, K_map, w_id):
    """
    Print de initiële kwalificatie (h[w,k,0]) en trainingsprogressie (t[w,k,0]) van een werknemer.
    """
    print(f"\n🧠 Debug: initiële kwalificaties van werknemer {w_id}")
    h = parsed_solution.get("h", {})
    t = parsed_solution.get("t", {})

    if not h:
        print("❌ Geen h-status gevonden in parsed_solution.")
        return

    for (w, k, s), val in h.items():
        if w == w_id and s == 0:
            t_val = t.get((w, k, s), 0)
            skill_name = K_map.get(k, f"Skill {k}")
            qualified = "✅" if val > 0.5 else "❌"
            print(f"  - {skill_name:<25}: h={val:.1f}   t={t_val:.1f}  {qualified}")

def debug_competency_matrix(P_map, K_map, r_p_k):
    """
    Print de vereiste skills per positie volgens de competentiematrix (r_p_k).
    """
    print("\n📋 Vereiste skills per positie volgens de competentiematrix:")
    pos_to_skills = {}

    for (p, k), val in r_p_k.items():
        if val != 1:
            continue
        pos_to_skills.setdefault(p, []).append(k)

    for p in sorted(pos_to_skills):
        pos_name = P_map.get(p, f"Pos {p}")
        skills = [K_map.get(k, f"Skill {k}") for k in sorted(pos_to_skills[p])]
        print(f"  - {pos_name:<30}: {', '.join(skills)}")
def calculate_team_flexibility_parsed(parsed_solution, Pm, P_map, q, f_w_s):

    """
    Bereken de Team Flexibility Coefficient (TVC) voor de eerste shift.

    Parameters:
    - parsed_solution: dict met 'x', 'q', ...
    - Pm: lijst van verplichte posities
    - P_map: mapping van positie-id naar naam
    - q: dict met kwalificaties (q[w, p, s])

ng empty TFC results
    Returns:
    - float: TVC in percentage
    """
    from HopcroftKarp import HopcroftKarp

    # 1. Bepaal de eerste shift waarop geplande inzet bestaat
    s_values = [s for (_, _, s) in parsed_solution.get("x", {})]
    if not s_values:
        return 0.0
    s_check = min(s_values)

    # 2. Posities die effectief ingevuld worden op s_check
    positions_to_fill = [
        p for p in Pm
        if any(parsed_solution["x"].get((w, p, s_check), 0) > 0.5 for w in f_w_s)
    ]
    valid_rights = set(positions_to_fill)

    # 3. Werknemers met geldige kwalificatie op s_check
    edge_list = [
        (w, p) for (w, p, s), val in q.items()
        if s == s_check and val >= 0.5 and p in valid_rights
    ]

    lefts = sorted(set(w for (w, _) in edge_list))
    rights = sorted(valid_rights)
    edge_dict = HopcroftKarp.build_edge_dict(edge_list)


    matches = HopcroftKarp.hopcroft_karp(lefts, rights, edge_dict)
    tvc = (len(matches) / len(rights)) * 100 if rights else 0.0

    return round(tvc, 1)



def debug_inzetbaarheid(f_w_s: dict, S: list, Pm: list, min_positions: int):
    """
    Debug de inzetbaarheid per shift.
    """

    results = []

    for s in S:
        inzetbare_workers = 0

        for w in f_w_s:
            if f_w_s[w][s] == 0:
                inzetbare_workers += 1

        result = {
            "shift": s,
            "available_workers": inzetbare_workers,
            "positions_needed": len(Pm),
            "feasible": inzetbare_workers >= len(Pm)
        }
        results.append(result)

    df = pd.DataFrame(results)

    print(f"\n📋 Shift inzetbaarheid-analyse: minimaal {len(Pm)} inzetbare werknemers vereist per shift.\n")

    violating = df[~df["feasible"]]
    if violating.empty:
        print("✅ Geen enkel capaciteitsprobleem gevonden.")
    else:
        print(f"⚠️ {len(violating)} shifts met tekort!\n")
        print(violating[["shift", "available_workers", "positions_needed"]])

    return df


def debug_inputs(inputs: dict):
    """
    Voert een uitgebreide debuganalyse op de inputs:
    - Skills zonder training load
    - Werknemers zonder mogelijkheid om vereiste skills te leren
    - Vereiste skills zonder beschikbare trainers

    Parameters:
    - inputs: de dictionary zoals teruggegeven door load_all_inputs of prepare_model_inputs
    """

    W = inputs["W"]
    K = inputs["K"]
    Pm = inputs["Pm"]
    r_p_k = inputs["r_p_k"]
    I_wks = inputs["I_wks"]
    L_k = inputs["L_k"]
    K_p = inputs["K_p"]
    P_k = inputs["P_k"]
    P_map = inputs["P_map"]
    K_map = inputs["K_map"]

    print("\n📋 START INPUT DEBUG")

    # 1. Zijn er skills zonder trainingsload?
    missing_Lk = [k for k in K if k not in L_k or L_k[k] <= 0]
    if missing_Lk:
        print(f"⚠️ Skills zonder geldige L_k: {[K_map[k] for k in missing_Lk]}")
    else:
        print("✅ Alle skills hebben een geldige trainingsload (L_k).")

    # 2. Kunnen werknemers überhaupt hun vereiste skills bereiken?
    for w in W:
        for p in Pm:
            required_skills = K_p.get(p, [])
            if not required_skills:
                continue
            for k in required_skills:
                can_train = (k in I_wks.get(w, {})) and (I_wks[w][k].get(0, 0) >= 0)
                if not can_train:
                    print(f"❌ Werknemer {w} kan vereiste skill '{K_map[k]}' (voor positie '{P_map[p]}') niet trainen.")

    # 3. Zijn er skills vereist waarvoor niemand kan trainen?
    for k in K:
        possible_trainers = sum(1 for w in W if (k in I_wks.get(w, {})) and (I_wks[w][k].get(0, 0) >= 0))
        if possible_trainers == 0:
            print(f"❌ Skill '{K_map[k]}' vereist, maar geen enkele werknemer kan trainen of heeft startniveau.")

    print("✅ Input debug volledig uitgevoerd.\n")
import pandas as pd

def export_inputs_to_excel(inputs: dict, export_path: str):
    """
    Exporteert alle relevante inputdata naar een Excelbestand.

    Parameters:
    - inputs: dictionary van inputdata
    - export_path: pad naar de Excel file (.xlsx)
    """

    # 1. Werknemerslijst
    df_w = pd.DataFrame({"EmployeeID": inputs["W"]})

    # 2. Skillslijst
    df_k = pd.DataFrame({
        "SkillID": list(inputs["K_map"].keys()),
        "SkillName": list(inputs["K_map"].values())
    })

    # 3. Positielijst
    df_p = pd.DataFrame({
        "PositionID": list(inputs["P_map"].keys()),
        "PositionName": list(inputs["P_map"].values())
    })

    # 4. Vereiste skills per positie (r_p_k)
    rpk_records = []
    for (p, k), val in inputs["r_p_k"].items():
        rpk_records.append({
            "PositionID": p,
            "PositionName": inputs["P_map"].get(p, "UNKNOWN"),
            "SkillID": k,
            "SkillName": inputs["K_map"].get(k, "UNKNOWN"),
            "Required": val
        })
    df_rpk = pd.DataFrame(rpk_records)

    # 5. Startskills per werknemer (I_wks)
    i_records = []
    for w in inputs["I_wks"]:
        for k in inputs["I_wks"][w]:
            for s in inputs["I_wks"][w][k]:
                i_records.append({
                    "EmployeeID": w,
                    "SkillID": k,
                    "Shift": s,
                    "CompletedHours": inputs["I_wks"][w][k][s]
                })
    df_iwks = pd.DataFrame(i_records)

    # 6. Experts
    e_records = []
    for w in inputs["experts"]:
        for k in inputs["experts"][w]:
            e_records.append({
                "EmployeeID": w,
                "SkillID": k,
                "IsExpert": inputs["experts"][w][k]
            })
    df_experts = pd.DataFrame(e_records)

    # 7. Training load (L_k)
    df_lk = pd.DataFrame([
        {"SkillID": k, "TrainingHours": v}
        for k, v in inputs["L_k"].items()
    ])

    # 8. Shift-vrije dagen (f_w_s)
    f_records = []
    for w in inputs["f_w_s"]:
        for s in inputs["f_w_s"][w]:
            f_records.append({
                "EmployeeID": w,
                "Shift": s,
                "OnLeave": inputs["f_w_s"][w][s]
            })
    df_fws = pd.DataFrame(f_records)

    # Exporteer naar Excel
    with pd.ExcelWriter(export_path) as writer:
        df_w.to_excel(writer, sheet_name="Employees", index=False)
        df_k.to_excel(writer, sheet_name="Skills", index=False)
        df_p.to_excel(writer, sheet_name="Positions", index=False)
        df_rpk.to_excel(writer, sheet_name="RequiredSkills", index=False)
        df_iwks.to_excel(writer, sheet_name="TrainingProgress", index=False)
        df_experts.to_excel(writer, sheet_name="Experts", index=False)
        df_lk.to_excel(writer, sheet_name="TrainingLoad", index=False)
        df_fws.to_excel(writer, sheet_name="VacationMatrix", index=False)

    print(f"✅ Inputgegevens succesvol geëxporteerd naar {export_path}")
def load_solution_from_excel(filepath: str) -> dict:
    """
    Laad een parsed oplossing uit een Excelbestand (zoals gegenereerd met export_solution_to_excel).
    """
    import pandas as pd

    parsed = {}
    xls = pd.ExcelFile(filepath)
    for sheet_name in xls.sheet_names:
        df = xls.parse(sheet_name)
        if "Value" not in df.columns:
            continue
        idx_cols = [c for c in df.columns if c.startswith("Index")]
        parsed[sheet_name] = {
            tuple(df.loc[i, idx_cols].astype(int)): df.loc[i, "Value"]
            for i in df.index
            if df.loc[i, "Value"] != 0
        }
    return parsed


def calculate_tfc_at_timepoints(parsed_solution, P_map, K_p, S, Pm, f_w_s, W):
    shifts_to_check = [90, 135, 180]
    result = {}

    # Check if required keys exist in parsed_solution
    required_keys = ["x", "y", "q"]
    if not all(key in parsed_solution for key in required_keys):
        raise ValueError("parsed_solution is missing required keys")

    for shift in shifts_to_check:
        # Skip if shift has no assignments
        if not any((w, p, shift) in parsed_solution["x"] for w in W for p in Pm):
            continue

        # Get the workers and their status for this shift
        workers_at_shift = []
        for w in W:
            # Check worker status
            if any(parsed_solution["x"].get((w, p, shift), 0) > 0.5 for p in Pm):
                status = "working"
            elif any(parsed_solution["y"].get((w, p, shift), 0) > 0.5 for p in Pm):
                status = "training"
            elif f_w_s.get((w, shift), 0) == 1:
                status = "leave"
            else:
                status = "reserve"
            workers_at_shift.append((w, status))

        # Get positions that need to be filled this shift
        positions_to_fill = [
            p for p in Pm
            if any(parsed_solution["x"].get((w, p, shift), 0) > 0.5
                   for w in W)
        ]

        # Determine qualified and unqualified workers
        qualified_workers = []
        unqualified_workers = []

        for (w, status) in workers_at_shift:
            for p in positions_to_fill:
                if parsed_solution["q"].get((w, p, shift), 0) >= 0.99:
                    qualified_workers.append((w, p))
                else:
                    unqualified_workers.append((w, p))

        # Calculate TFC
        tfc = calculate_team_flexibility(
            qualified_workers=qualified_workers,
            unqualified_workers=unqualified_workers,
            positions_to_fill=positions_to_fill
        )

        result[shift] = {
            "qualified_workers": qualified_workers,
            "unqualified_workers": unqualified_workers,
            "tfc": tfc,
            "positions_to_fill": positions_to_fill
        }

    return result

from collections import defaultdict
import itertools
from typing import List, Tuple


def calculate_team_flexibility(qualified_workers: List[Tuple[str, str]],
                               unqualified_workers: List[Tuple[str, str]],
                               positions_to_fill: List[str]) -> float:
    """
    Calculate Team Flexibility Coefficient (TFC) based on worker qualifications and positions needed.

    Args:
        qualified_workers: List of (worker_id, position) tuples for qualified workers
        unqualified_workers: List of (worker_id, position) tuples for unqualified workers
        positions_to_fill: List of position names that need to be filled

    Returns:
        TFC percentage (0-100) indicating team flexibility
    """
    # Combine all unique workers (both qualified and unqualified)
    all_workers = {w for w, _ in qualified_workers} | {w for w, _ in unqualified_workers}
    n_positions = len(positions_to_fill)

    if not all_workers or n_positions == 0:
        return 0.0

    # Create a dictionary of each worker's qualified positions
    worker_qualifications = defaultdict(set)
    for w, p in qualified_workers:
        worker_qualifications[w].add(p)

    # Convert to list of TeamMember-like tuples (id, positions)
    team_members = [(w, worker_qualifications[w]) for w in all_workers]

    # Generate all possible team combinations
    all_combinations = itertools.combinations(team_members, n_positions)
    weighted_total = 0.0
    weighted_nok = 0.0

    for combination in all_combinations:
        rights = set(positions_to_fill)
        edges = {
            w: {p for p in positions if p in rights}
            for w, positions in combination
        }

        # Perform bipartite matching (simplified version)
        matches = hopcroft_karp_bipartite_matching(set(edges.keys()), rights, edges)

        # All combinations are equally weighted in basic TFC
        weight = 1.0
        weighted_total += weight

        if len(matches) < n_positions:
            weighted_nok += weight

    # Calculate final TFC
    if weighted_total == 0:
        return 0.0
    return ((weighted_total - weighted_nok) / weighted_total) * 100.0


def hopcroft_karp_bipartite_matching(lefts, rights, edges):
    """Simplified bipartite matching algorithm (Hopcroft-Karp variant)"""
    to_right = {l: None for l in lefts}
    to_left = {r: None for r in rights}

    def bpm(u, visited):
        for v in edges.get(u, set()):
            if v not in visited:
                visited.add(v)
                if to_left[v] is None or bpm(to_left[v], visited):
                    to_left[v] = u
                    to_right[u] = v
                    return True
        return False

    for l in lefts:
        bpm(l, set())

    return {k: v for k, v in to_right.items() if v is not None}

def check_training_feasibility(inputs: dict):
    W = inputs["W"]
    K = inputs["K"]
    Pm = inputs["Pm"]
    K_p = inputs["K_p"]
    P_map = inputs["P_map"]
    K_map = inputs["K_map"]
    I_wks = inputs["I_wks"]
    experts = inputs["experts"]

    print("\n📋 CHECK TRAINING FEASIBILITY")

    # 1. Skills zonder expert
    print("\n❗ Skills zonder expert:")
    for k in K:
        is_expert_available = any(experts.get(w, {}).get(k, 0) == 1 for w in W)
        if not is_expert_available:
            print(f"  - Skill {k} ({K_map.get(k, f'Skill {k}')}): ❌ GEEN expert beschikbaar")

    # 2. Skills zonder trainbare werknemer
    print("\n❗ Skills die niemand kan trainen (I_wks):")
    for k in K:
        can_anyone_train = any(k in I_wks.get(w, {}) for w in W)
        if not can_anyone_train:
            print(f"  - Skill {k} ({K_map.get(k, f'Skill {k}')}): ❌ niemand heeft trainingsentry")

    # 3. Posities waarvan geen enkele skill trainbaar is
    print("\n❗ Posities die mogelijk onhaalbaar zijn (geen enkele skill kan worden opgebouwd):")
    for p in Pm:
        untrainable_skills = []
        for k in K_p.get(p, []):
            has_expert = any(experts.get(w, {}).get(k, 0) == 1 for w in W)
            has_train_entry = any(k in I_wks.get(w, {}) for w in W)
            if not has_expert or not has_train_entry:
                untrainable_skills.append(k)
        if untrainable_skills:
            pos_name = P_map.get(p, f"Pos {p}")
            print(f"  - {pos_name}: ❗ onhaalbare skills: {[K_map[k] for k in untrainable_skills]}")
        else:
            pass  # alle trainbaar

    print("\n✅ Check voltooid.\n")
import pandas as pd

def export_all_inputs_to_excel(inputs: dict, output_path: str = "output/all_inputs_export.xlsx"):
    """
    Exporteert alle belangrijke inputstructuren naar een Excelbestand met meerdere tabbladen.

    Parameters:
    - inputs: dictionary zoals gegenereerd door `load_all_inputs(...)`
    - output_path: pad naar het exportbestand
    """

    def safe_df(data, columns=None):
        return pd.DataFrame(data, columns=columns) if columns else pd.DataFrame(data)

    with pd.ExcelWriter(output_path) as writer:
        # Werknemers
        safe_df(inputs["W"], columns=["EmployeeID"]).to_excel(writer, sheet_name="Employees", index=False)

        # Skills
        pd.DataFrame([
            {"SkillID": k, "SkillName": v} for k, v in inputs["K_map"].items()
        ]).to_excel(writer, sheet_name="Skills", index=False)

        # Posities
        pd.DataFrame([
            {"PositionID": p, "PositionName": name} for p, name in inputs["P_map"].items()
        ]).to_excel(writer, sheet_name="Positions", index=False)

        # Vereiste skills per positie
        pd.DataFrame([
            {
                "PositionID": p,
                "PositionName": inputs["P_map"].get(p, f"Pos {p}"),
                "SkillID": k,
                "SkillName": inputs["K_map"].get(k, f"Skill {k}")
            }
            for (p, k), val in inputs["r_p_k"].items() if val == 1
        ]).to_excel(writer, sheet_name="RequiredSkills", index=False)

        # I_wks (training progress)
        pd.DataFrame([
            {"EmployeeID": w, "SkillID": k, "Shift": s, "Hours": v}
            for w in inputs["I_wks"]
            for k in inputs["I_wks"][w]
            for s, v in inputs["I_wks"][w][k].items()
        ]).to_excel(writer, sheet_name="I_wks_Training", index=False)

        # Experts
        pd.DataFrame([
            {"EmployeeID": w, "SkillID": k}
            for w in inputs["experts"]
            for k in inputs["experts"][w]
            if inputs["experts"][w][k] == 1
        ]).to_excel(writer, sheet_name="Experts", index=False)

        # Training load (L_k)
        pd.DataFrame([
            {"SkillID": k, "TrainingLoad": v}
            for k, v in inputs["L_k"].items()
        ]).to_excel(writer, sheet_name="TrainingLoad", index=False)

        # f_w_s (beschikbaarheid)
        pd.DataFrame([
            {"EmployeeID": w, "Shift": s, "OnLeave": val}
            for w in inputs["f_w_s"]
            for s, val in inputs["f_w_s"][w].items()
        ]).to_excel(writer, sheet_name="LeaveMatrix", index=False)

        # Alfa_p gewichten
        if "alfa_p" in inputs:
            pd.DataFrame([
                {"PositionID": p, "Weight": w}
                for p, w in inputs["alfa_p"].items()
            ]).to_excel(writer, sheet_name="AlfaWeights", index=False)
        # Skills per werknemer (alleen als uren > 0)
        pd.DataFrame([
            {
                "EmployeeID": w,
                "SkillID": k,
                "SkillName": inputs["K_map"].get(k, f"Skill {k}"),
                "HoursCompleted": inputs["I_wks"][w][k][0]
            }
            for w in inputs["I_wks"]
            for k in inputs["I_wks"][w]
            if inputs["I_wks"][w][k][0] > 0
        ]).to_excel(writer, sheet_name="EmployeeSkills", index=False)

        # Simulatieparameters
        if "sim_params" in inputs:
            pd.DataFrame([
                {"Month": m, "LeaveProb": p}
                for m, p in inputs["sim_params"].get("month_probs", {}).items()
            ]).to_excel(writer, sheet_name="LeaveProbabilities", index=False)

            pd.DataFrame([{
                "MeanLeaveDuration": inputs["sim_params"].get("verlofduur", None),
                "MutationProb": inputs["sim_params"].get("mutation_prob", None)
            }]).to_excel(writer, sheet_name="SimParams", index=False)


    print(f"✅ Alle inputgegevens geëxporteerd naar: {output_path}")
