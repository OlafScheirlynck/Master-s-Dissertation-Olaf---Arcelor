import itertools
# remove terugbrengen naar 11
#opleidingen met 0 niet aanbieden
import pandas as pd
from collections import defaultdict, deque
from openpyxl import load_workbook

# === Inlezen van competentiematrix en bepalen van inzetbaarheid per positie ===
competentiepad = r"C:\Users\olafs\OneDrive\Bureaublad\master's dissertation\variation and competences\AM_Data\SPL_SDG_competentiematrix_VOLLEDIG_stamnummers.xlsx"
trainingspad = r"C:\Users\olafs\OneDrive\Bureaublad\master's dissertation\variation and competences\AM_Data\Trainings.xlsx"
part_time_path = r"C:\Users\olafs\OneDrive\Bureaublad\master's dissertation\variation and competences\AM_Data\SDG 1 ABCD partttime 20250325.xlsx"
part_time_df = pd.read_excel(part_time_path, sheet_name="SDG 1 ABCD")

# Toegestane medewerkers per team uit inzetbaarheidsbestanden
inzetbaarheid_files = {
    "SDG 1 Ploeg A": "SDG 1 - Ploeg A - wn inzetbaarheden.xlsx",
    "SDG 1 Ploeg B": "SDG 1 - Ploeg B - wn inzetbaarheden.xlsx",
    "SDG 1 Ploeg C": "SDG 1 - Ploeg C - wn inzetbaarheden.xlsx",
    "SDG 1 Ploeg D": "SDG 1 - Ploeg D - wn inzetbaarheden.xlsx",
}

# Maak een dictionary: {employee_id: beschikbaarheid}
employee_availability_map = {
    int(row["Fictive EmployeeId"]): int(row["ProcentParttime"])
    for _, row in part_time_df.iterrows()
}
all_r_p_k = {}
all_K_map = {}
all_has_skill = {}
all_inzetbaarheid = {}
all_blocked_skills = {}

def load_training_data_dynamic(file_path, team_sheet, k_name_map, toegestane_ids):
    header_row = 0
    df = pd.read_excel(file_path, sheet_name=team_sheet, header=header_row)
    df.columns = [str(col).strip() for col in df.columns]

    if "Fictive EmployeeId" not in df.columns or "TaskDescription" not in df.columns:
        raise ValueError(f"Kolomnamen niet gevonden in {team_sheet}")

    worker_ids = sorted([int(w) for w in df["Fictive EmployeeId"].dropna().unique() if int(w) in toegestane_ids])
    W = list(worker_ids)
    has_skill = {w: {} for w in W}
    blocked_skills = {w: set() for w in W}

    normalized_map = {str(name).strip().lower(): k_idx for k_idx, name in k_name_map.items()}

    # Verzamel alle employability scores per (werknemer, skill)
    employability_tracker = defaultdict(list)

    for _, row in df.iterrows():
        if pd.isna(row["Fictive EmployeeId"]) or pd.isna(row["TaskDescription"]):
            continue
        w = int(row["Fictive EmployeeId"])
        if w not in toegestane_ids:
            continue
        k = str(row["TaskDescription"]).strip().lower()
        if k not in normalized_map:
            continue
        k_idx = normalized_map[k]
        employability = int(row.get("Fictive EmployeeId.1", row.get("Fictive EmployeeId", 0)))
        employability_tracker[(w, k_idx)].append(employability)

    # Zet skills pas op True als géén enkele score 0 is, én minstens één score ≥ 2
    for (w, k_idx), scores in employability_tracker.items():
        if 0 in scores:
            blocked_skills[w].add(k_idx)
            continue  # uitgesloten
        if any(e >= 2 for e in scores):
            has_skill[w][k_idx] = True

    return W, has_skill, blocked_skills

wb = load_workbook(competentiepad, data_only=True)
data_rows_by_sheet = {sheet.title: list(sheet.iter_rows(values_only=True)) for sheet in wb.worksheets}

team_mapping = {
    "SDG 1 Ploeg A": "PLOEG A SDG1",
    "SDG 1 Ploeg B": "PLOEG B SDG1",
    "SDG 1 Ploeg C": "PLOEG C SDG1",
    "SDG 1 Ploeg D": "PLOEG D SDG1",
}

all_teams_positions = {}


def get_toegestane_ids(ws_name):
    inzetbaarheid_folder = r"C:\Users\olafs\OneDrive\Bureaublad\master's dissertation\variation and competences\AM_Data\RE_ extra data taken"
    inzetbaarheid_path = f"{inzetbaarheid_folder}\\{inzetbaarheid_files[ws_name]}"

    df = pd.read_excel(inzetbaarheid_path)

    # Filter enkel rijen waar "Gebruik In Telling Kal." gelijk is aan 1.0 of True
    df = df[df["Gebruik In Telling Kal."].isin([1, 1.0, True, "True"])]

    # Verwijder rijen zonder geldig stamnummer
    df = df[pd.notna(df["Fictief Stam Nr."])]

    # Zet om naar int
    return set(df["Fictief Stam Nr."].astype(int))


for ws_name, training_sheet in team_mapping.items():
    print(f"\n🧩 Verwerking van team: {ws_name}")
    try:
        toegestane_ids = get_toegestane_ids(ws_name)
        data_rows = data_rows_by_sheet[ws_name]

        # Zoek locatie van ### TAKEN ###
        taken_start_row = next(i for i, row in enumerate(data_rows) if any(cell == "### TAKEN ###" for cell in row))
        taken_row = data_rows[taken_start_row]
        taken_start_col = next(i for i, val in enumerate(taken_row) if val == "### TAKEN ###") + 1

        K_names = [str(val).strip() for val in taken_row[taken_start_col:] if isinstance(val, str) and val.strip()]
        K = list(range(len(K_names)))
        K_map = dict(zip(K, K_names))
        K_reverse_map = {v: k for k, v in K_map.items()}

        header_row_index = next(i for i, row in enumerate(data_rows) if "Positie Naam" in row)
        header_row = data_rows[header_row_index]
        positie_col = header_row.index("Positie Naam")
        verplicht_col = header_row.index(next(c for c in header_row if isinstance(c, str) and "Verplicht" in c))

        verplichte_posities = []
        for row in data_rows[header_row_index + 1:]:
            if row[positie_col] and str(row[verplicht_col]).strip() == "1":
                verplichte_posities.append(str(row[positie_col]).strip())

        P_names = verplichte_posities
        P = list(range(len(P_names)))
        P_map = dict(zip(P, P_names))
        P_reverse_map = {v: k for k, v in P_map.items()}

        r_p_k = {}
        matrix_start_index = taken_start_row + 3
        for row in data_rows[matrix_start_index:]:
            positie_naam = row[positie_col]
            if not isinstance(positie_naam, str) or not positie_naam.strip():
                continue
            verplicht = str(row[verplicht_col]).strip() if verplicht_col < len(row) else "0"
            if verplicht != "1":
                continue
            p_index = P_reverse_map.get(positie_naam)
            for i_k, taaknaam in enumerate(K_names):
                col_idx = taken_start_col + i_k
                if col_idx < len(row) and row[col_idx] == 1:
                    r_p_k[(p_index, i_k)] = 1

        all_r_p_k[ws_name] = r_p_k
        all_K_map[ws_name] = K_map

        W, has_skill, blocked_skills = load_training_data_dynamic(trainingspad, training_sheet, K_map, toegestane_ids)
        all_has_skill[ws_name] = has_skill
        all_blocked_skills[ws_name] = blocked_skills

        ploeg_structuur = defaultdict(set)
        for p in P:
            vereist = [k for (p_check, k) in r_p_k if p_check == p]
            for w in W:
                skills = has_skill.get(w, {})
                if all(skills.get(k, False) for k in vereist):
                    ploeg_structuur[P_map[p]].add(w)
        # Zorg dat ook niet-inzetbare werknemers zichtbaar zijn
        alle_ids_in_structuur = set()
        for mensen in ploeg_structuur.values():
            alle_ids_in_structuur.update(mensen)

        for w in W:
            if w not in alle_ids_in_structuur:
                # Voeg toe aan ploegstructuur met lege set (geen inzetbare posities)
                ploeg_structuur["Geen inzetbare posities"].add(w)
        inzetbaarheid = defaultdict(list)
        for pos, mensen in ploeg_structuur.items():
            for w in mensen:
                inzetbaarheid[w].append(pos)
        all_inzetbaarheid[ws_name] = inzetbaarheid

        all_teams_positions[ws_name] = ploeg_structuur

    except Exception as e:
        print(f"⚠️ Fout bij team {ws_name}: {e}")




        #W, has_skill = load_training_data_dynamic(trainingspad, training_sheet, K_map)

        # === Bepaal inzetbaarheid
        ploeg_structuur = defaultdict(set)
        for p in P:
            vereist = [k for (p_check, k) in r_p_k if p_check == p]
            for w in W:
                skills = has_skill.get(w, {})
                if all(skills.get(k, False) for k in vereist):
                    ploeg_structuur[P_map[p]].add(w)

        all_teams_positions[ws_name] = ploeg_structuur


# === 2. Klassen ===
class EmployabilityCalculationManager:
    def __init__(self, team): self.team = team

    def calculate(self):
        employee_results = {m.employee_id: EmployeeResult(m.employee_id, m.employee_positions) for m in self.team.team_members}
        team_result = TeamResult(self.team.team_id, self.team.positions_to_fill)
        n = len(self.team.positions_to_fill)
        if len(self.team.team_members) < n:
            print("⚠️ Niet genoeg mensen.")
            return list(employee_results.values()), team_result

        all_combinations = itertools.combinations(self.team.team_members, n)
        weighted_total = 0.0
        weighted_nok = 0.0
        total_available = sum(m.availability for m in self.team.team_members)

        for drawn_team in all_combinations:
            rights = set(self.team.positions_to_fill)
            edges = {
                m.employee_id: set(p for p in m.employee_positions if p in self.team.positions_to_fill)
                for m in drawn_team
            }

            lefts = set(edges.keys())
            matches = EmployabilityCalculationManager.hopcroft_karp(lefts, rights, edges)

            total_team_availability = sum(m.availability for m in drawn_team)
            team_availability_factor = total_team_availability / total_available if total_available != 0 else 0

            weighted_total += team_availability_factor
            if len(matches) != n:
                weighted_nok += team_availability_factor
                for key in team_result.missing_functions.keys():
                    if key not in matches.values():
                        team_result.missing_functions[key] += 1
                        break
            for employee_id, result in employee_results.items():
                if employee_id not in edges:
                    result.result_positions["Rest"] += 1
                    if len(matches) != n:
                        result.result_absent_nok += 1 * team_availability_factor  # Adjust by availability factor
                    else:
                        result.result_absent_ok += 1 * team_availability_factor
                elif employee_id not in matches:
                    result.result_positions["Unused"] += 1
                    result.result_present_nok += 1 * team_availability_factor  # Adjust by availability factor
                else:
                    pos = matches[employee_id]
                    result.result_positions[pos] += 1
                    if len(matches) != n:
                        result.result_present_nok += 1 * team_availability_factor  # Adjust by availability factor
                        if pos not in edges[employee_id]:
                            result.result_underqualified[pos] += 1
                            team_result.underqualified_assignments[pos] += 1
                    else:
                        result.result_present_ok += 1 * team_availability_factor  # Adjust by availability factor

        team_result.flexibility = (weighted_total - weighted_nok) / weighted_total * 100.0 if weighted_total > 0 else 0

        return list(employee_results.values()), team_result

    @staticmethod
    def has_augmenting_path(lefts, edges, to_r, to_l, d):
        q = deque()
        for l in lefts:
            d[l] = 0 if to_r[l] == "" else float("inf")
            if to_r[l] == "": q.append(l)
        d[""] = float("inf")
        while q:
            l = q.popleft()
            if d[l] < d[""]:
                for r in edges[l]:
                    nl = to_l[r]
                    if d[nl] == float("inf"):
                        d[nl] = d[l] + 1
                        q.append(nl)
        return d[""] != float("inf")

    @staticmethod
    def try_matching(l, edges, to_r, to_l, d):
        if l == "": return True
        for r in edges[l]:
            nl = to_l[r]
            if d[nl] == d[l] + 1 and EmployabilityCalculationManager.try_matching(nl, edges, to_r, to_l, d):
                to_l[r] = l
                to_r[l] = r
                return True
        d[l] = float("inf")
        return False

    @staticmethod
    def hopcroft_karp(lefts, rights, edges):
        d = {}
        to_r = {l: "" for l in lefts}
        to_l = {r: "" for r in rights}
        while EmployabilityCalculationManager.has_augmenting_path(lefts, edges, to_r, to_l, d):
            for l in (l for l in lefts if to_r[l] == ""):
                EmployabilityCalculationManager.try_matching(l, edges, to_r, to_l, d)
        return {k: v for k, v in to_r.items() if v != ""}

class EmployeeResult:
    def __init__(self, id, pos):
        self.employee_id = id
        self.employee_positions = pos
        self.result_positions = defaultdict(int)
        self.result_present_ok = 0
        self.result_present_nok = 0
        self.result_absent_ok = 0
        self.result_absent_nok = 0
        self.result_underqualified = defaultdict(int)

class TeamResult:
    def __init__(self, id, positions):
        self.team_id = id
        self.positions_to_fill = positions
        self.missing_functions = defaultdict(int)
        self.flexibility = 0.0
        self.underqualified_assignments = defaultdict(int)

class TeamMember:
    def __init__(self, id, avail, pos, part_time_percentage=100):
        self.employee_id = id
        self.availability = avail * (part_time_percentage / 100)  # Adjust availability based on part-time percentage
        self.employee_positions = pos
class Team:
    def __init__(self, id, members, pos): self.team_id = id; self.team_members = members; self.positions_to_fill = pos

# === 3. Sensitivity Analysis functie ===
def perform_sensitivity_analysis_on_team(team_name: str):
    inzetbaarheid = all_inzetbaarheid[team_name]
    K_map = all_K_map[team_name]
    has_skill = all_has_skill[team_name]

    if team_name not in all_teams_positions:
        print(f"⚠️ Team '{team_name}' niet gevonden.")
        return


    pos_map = all_teams_positions[team_name]
    all_people = set()
    for pos, names in pos_map.items():
        all_people.update(names)

    positions = [p for p in all_teams_positions[gekozen_team].keys() if p in P_map.values()]


    # Analyse zonder dummy's
    members_no_dummy = []
    for name in all_people:
        part_time_percentage = employee_availability_map.get(int(name), 100)
        members_no_dummy.append(TeamMember(name, part_time_percentage, inzetbaarheid[name]))

    base_team = Team(team_name + " (original)", members_no_dummy, positions)
    df_no_dummy = run_analysis(base_team, positions)

    # Analyse met dummy's
    members_with_dummy = add_dummy_members_to_team(members_no_dummy, positions, target_size=12)
    team_with_dummies = Team(team_name + " (with dummy)", members_with_dummy, positions)
    df_with_dummy = run_analysis(team_with_dummies, positions)
    debug_rows = []
    for m in team_with_dummies.team_members:
        is_dummy = str(m.employee_id).startswith("DUMMY")
        if is_dummy:
            skills = "—"
            positions = "—"
        else:
            # Zet skillnamen en posities als strings
            skills = ", ".join(
                sorted(K_map[k] for k in has_skill.get(int(m.employee_id), {}) if has_skill[int(m.employee_id)].get(k)))
            positions = ", ".join(sorted(m.employee_positions))
        debug_rows.append({
            "Employee": m.employee_id,
            "Availability": m.availability,
            "Skills": skills,
            "Positions": positions
        })
    debug_df = pd.DataFrame(debug_rows)
    # Schrijf naar Excel met twee sheets
    output_path = f"Sensitivity_{team_name.replace(' ', '_')}.xlsx"
    with pd.ExcelWriter(output_path) as writer:
        df_no_dummy.to_excel(writer, index=False, sheet_name="Original")
        df_with_dummy.to_excel(writer, index=False, sheet_name="With_Dummy")
        df_strategisch.to_excel(writer, index=False, sheet_name="Strategic_Selection")
        debug_df.to_excel(writer, index=False, sheet_name="Skills_and_Positions")

    print(f"✅ Resultaten opgeslagen in: {output_path}")
    return df_no_dummy, df_with_dummy

def add_dummy_members_to_team(members, positions_to_fill, target_size=15):
    """
    Vul team aan tot target_size met werknemers zonder inzetbare posities.
    """
    current_ids = {m.employee_id for m in members}
    dummies = []
    idx = 1
    while len(members) + len(dummies) < target_size:
        dummy_id = f"DUMMY_{idx}"
        while dummy_id in current_ids:
            idx += 1
            dummy_id = f"DUMMY_{idx}"
        dummies.append(TeamMember(dummy_id, 100, []))  # geen inzetbare posities
        idx += 1
    return members + dummies
def run_analysis(team: Team, positions: list) -> pd.DataFrame:
    base_manager = EmployabilityCalculationManager(team)
    _, base_result = base_manager.calculate()
    base_tfc = base_result.flexibility

    print(f"📊 Sensitivity analysis for team '{team.team_id}' (TFC: {base_tfc:.2f}%)")
    rows = []

    # Extract base name (e.g. "SDG 1 Ploeg A") from team.team_id
    base_team_name = team.team_id.split(" (")[0]
    blocked_skills_for_team = all_blocked_skills.get(base_team_name, {})

    for m in team.team_members:
        # 1️⃣ Dummy vervangt werknemer
        red_team = []
        for p in team.team_members:
            if p.employee_id == m.employee_id:
                dummy_id = f"DUMMY_FOR_{m.employee_id}"
                red_team.append(TeamMember(dummy_id, 100, []))  # dummy zonder posities
            else:
                red_team.append(p)
        test_team = Team(team.team_id, red_team, positions)
        _, res = EmployabilityCalculationManager(test_team).calculate()
        rows.append({
            "Employee": m.employee_id,
            "Change Type": "Removed from team (dummy filled)",
            "Test Position": "",
            "New TFC": res.flexibility,
            "Delta TFC": res.flexibility - base_tfc
        })

        # 2️⃣ Trainable positions: alleen als persoon GEEN 0-score had op vereiste skills
        all_pos = set(positions)
        rpk = all_r_p_k[base_team_name]
        trainable_pos = allowed_training_positions(
            m.employee_positions, all_pos, m.employee_id, rpk, blocked_skills_for_team
        )

        for add_pos in trainable_pos:
            temp = TeamMember(m.employee_id, m.availability, m.employee_positions + [add_pos])
            new_team = [temp if p.employee_id == m.employee_id else p for p in team.team_members]
            test_team = Team(team.team_id, new_team, positions)
            _, res = EmployabilityCalculationManager(test_team).calculate()
            rows.append({
                "Employee": m.employee_id,
                "Change Type": "Gain position",
                "Test Position": add_pos,
                "New TFC": res.flexibility,
                "Delta TFC": res.flexibility - base_tfc
            })

        # 3️⃣ Posities verliezen, als persoon meerdere posities heeft
        if len(m.employee_positions) > 1:
            for remove_pos in m.employee_positions:
                temp = TeamMember(m.employee_id, m.availability, [p for p in m.employee_positions if p != remove_pos])
                new_team = [temp if p.employee_id == m.employee_id else p for p in team.team_members]
                test_team = Team(team.team_id, new_team, positions)
                _, res = EmployabilityCalculationManager(test_team).calculate()
                rows.append({
                    "Employee": m.employee_id,
                    "Change Type": "Lose position",
                    "Test Position": remove_pos,
                    "New TFC": res.flexibility,
                    "Delta TFC": res.flexibility - base_tfc
                })

    return pd.DataFrame(rows)


'''
def allowed_training_positions(current_positions, all_positions):
    current = set(current_positions)
    possible = []

    # Regel 1: GA als je nog niets hebt
    if not current:
        if "GA" in all_positions:
            possible.append("GA")

    # Regel 2: IN en UIT enkel als je alleen GA hebt
    if {"GA"} in current:
        for pos in ["IN", "UIT"]:
            if pos in all_positions:
                possible.append(pos)

    # Regel 3: INS als je IN en UIT hebt
    if {"ING", "UIT"} in current:
        if "INS" in all_positions and "INS" not in current:
            possible.append("INS")

    # Regel 4: CEN en AFL pas als je ING, UIT en INS hebt
    if {"ING", "UIT", "INS"} in current:
        for pos in ["CEN", "AFL"]:
            if pos in all_positions and pos not in current:
                possible.append(pos)

        # Regel 5: PLB pas als je CEN hebt
    if "CEN" in current and "PLB" in all_positions and "PLB" not in current:
            possible.append("PLB")

    return possible
    '''
def allowed_training_positions(current_positions, all_positions, employee_id, r_p_k, blocked_skills):
    """
    Simpele versie: iemand mag elke positie 'trainen' die hij nog niet kan.
    """
    current = set(current_positions)
    blocked = blocked_skills.get(employee_id, set())
    allowed = []
    for pos in all_positions - current:
        p_idx = P_reverse_map.get(pos)
        if p_idx is None:
            continue
        vereiste_skills = [k for (p, k) in r_p_k if p == p_idx]
        if all(k not in blocked for k in vereiste_skills):
            allowed.append(pos)
    return allowed
    return list(all_positions - set(current_positions))


def print_team_debug_info(team_name: str):
    print(f"\n📋 DEBUG INFO VOOR TEAM: {team_name}")

    if team_name not in all_teams_positions:
        print("⚠️ Team niet gevonden.")
        return

    pos_map = all_teams_positions[team_name]
    alle_posities = list(P_map.values())  # alle gedefinieerde verplichte posities
    print(f"🔹 Aantal posities: {len(alle_posities)}")
    print(f"🔹 Posities: {alle_posities}")

    # Lijst alle werknemers en hun inzetbaarheid
    inzetbaarheid = defaultdict(list)
    for pos, mensen in ploeg_structuur.items():
        for w in mensen:
            inzetbaarheid[w].append(pos)
    all_inzetbaarheid[ws_name] = inzetbaarheid

    for pos, mensen in pos_map.items():
        for m in mensen:
            inzetbaarheid[m].append(pos)

    print(f"\n👷 Werknemers en hun inzetbare posities:")
    for w, posities in inzetbaarheid.items():
        print(f"- Werknemer {w}: {', '.join(posities)}")

    # Optioneel: toon ook vereiste skills per positie
    if "r_p_k" in globals():
        print(f"\n📌 Vereiste skills per positie:")
        pos_skill_map = defaultdict(list)
        for (p_idx, k_idx), val in r_p_k.items():
            if val == 1:
                pos_skill_map[P_map[p_idx]].append(K_map[k_idx])
        for p, skills in pos_skill_map.items():
            print(f"- {p}: {', '.join(skills)}")


def validate_position_mapping(team_name: str):
    print(f"\n🛠️ Validatie van mappings voor team '{team_name}'")

    if team_name not in all_teams_positions:
        print("⚠️ Team niet gevonden.")
        return

    pos_map = all_teams_positions[team_name]
    pos_inzetbaar = set(pos_map.keys())
    pos_in_rpk = set(P_map[p] for (p, _) in r_p_k.keys())
    pos_in_pmap = set(P_map.values())

    print("\n🔍 Posities met skills in r_p_k maar NIET in ploegstructuur:")
    for p in sorted(pos_in_rpk - pos_inzetbaar):
        print(f"⚠️ {p}")

    print("\n🔍 Posities in ploegstructuur zonder skills in r_p_k:")
    for p in sorted(pos_inzetbaar - pos_in_rpk):
        print(f"⚠️ {p}")

    print("\n🔍 Posities die wel skills hebben maar GEEN inzetbare mensen:")
    for p in sorted(pos_in_rpk):
        if p in pos_map and len(pos_map[p]) == 0:
            print(f"⚠️ {p} heeft vereisten maar geen inzetbare mensen.")

    print("\n🔍 Posities met inzetbare mensen maar GEEN skillvereisten:")
    for p in sorted(pos_inzetbaar):
        if p not in pos_in_rpk:
            print(f"⚠️ {p} heeft mensen, maar geen skillvereisten in r_p_k.")

    print("\n✅ Mappingcontrole voltooid.")


# === 4. Run voor gekozen team ===




print("\nBeschikbare teams:")
for team in all_teams_positions.keys():
    print("-", team)

gekozen_team = input("\nWelke ploeg wil je analyseren? (exacte naam uit bovenstaande lijst): ").strip()
validate_position_mapping(gekozen_team)
print_team_debug_info(gekozen_team)

# Na de standaard run van perform_sensitivity_analysis_on_team
positions = [p for p in all_teams_positions[gekozen_team].keys() if p != "Geen inzetbare posities"]
inzetbaarheid = defaultdict(list)
for pos, mensen in all_teams_positions[gekozen_team].items():
    for m in mensen:
        inzetbaarheid[m].append(pos)

# Selectie op basis van inzetbaarheid
import math
target_team_size = math.ceil(len(positions) * 1.4 + 1)
gesorteerd = sorted(inzetbaarheid.items(), key=lambda x: (-len(x[1]), x[0]))
geselecteerde_mensen = []
for werker, posities_werker in gesorteerd:
    geselecteerde_mensen.append(werker)
    if len(geselecteerde_mensen) >= target_team_size:
        break

members_140 = []
# Adjust TeamMember creation to handle part-time availability
for name in geselecteerde_mensen:
    part_time_percentage = employee_availability_map.get(int(name), 100)
    members_140.append(TeamMember(name, part_time_percentage, inzetbaarheid[name]))  # Create team member with adjusted availability


# Create Team with adjusted availability
selected_team = Team(f"{gekozen_team} (strategic sample)", members_140, positions)

# Now proceed with the analysis using this adjusted team
df_strategisch = run_analysis(selected_team, positions)

# Save the analysis results to Excel
with pd.ExcelWriter(f"Sensitivity_{gekozen_team.replace(' ', '_')}.xlsx", mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    df_strategisch.to_excel(writer, index=False, sheet_name="Strategic_Selection")

perform_sensitivity_analysis_on_team(gekozen_team)