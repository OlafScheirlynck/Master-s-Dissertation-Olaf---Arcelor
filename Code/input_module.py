from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
import os
import numpy as np
from scipy.stats import gaussian_kde
import math
from Code.reality_module import simulate_vacation_days

class EmpiricalSampler:
    def __init__(self, data, use_kde=False, bandwidth_factor=1.0):
        self.data = np.array(data)
        self.use_kde = use_kde

        if self.use_kde:
            self.kde = gaussian_kde(self.data)
            self.kde.set_bandwidth(self.kde.factor * bandwidth_factor)
        else:
            self.kde = None

    def sample(self):
        if self.use_kde and self.kde is not None:
            sample = self.kde.resample(1)[0][0]
            sample = max(1, int(round(sample)))
        else:
            sample = np.random.choice(self.data)
        return sample

def load_competence_structure(filepath: str, sheet_name: str) -> dict:
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]
    data_rows = list(ws.iter_rows(values_only=True))

    taken_start_row = next(i for i, row in enumerate(data_rows) if any(cell == "### TAKEN ###" for cell in row))
    taken_row = data_rows[taken_start_row]
    taken_start_col = next(i for i, val in enumerate(taken_row) if val == "### TAKEN ###") + 1

    K_names = [str(val).strip().upper() for val in taken_row[taken_start_col:] if isinstance(val, str) and val.strip()]
    K = list(range(len(K_names)))
    K_map = dict(zip(K, K_names))
    K_reverse_map = {v: k for k, v in K_map.items()}

    header_row_index = next(i for i, row in enumerate(data_rows) if "Positie Naam" in row)
    header_row = data_rows[header_row_index]
    positie_col = header_row.index("Positie Naam")
    verplicht_col = header_row.index(next(c for c in header_row if isinstance(c, str) and "Verplicht" in c))

    verplichte_posities = [str(row[positie_col]).strip()
                           for row in data_rows[header_row_index + 1:]
                           if row[positie_col] and str(row[verplicht_col]).strip() == "1"]

    trainingsposities = [p + " training" for p in verplichte_posities]
    reserveposities = ["reserve"]
    P_names = verplichte_posities + trainingsposities + reserveposities
    P = list(range(len(P_names)))
    P_map = dict(zip(P, P_names))
    P_reverse_map = {v: k for k, v in P_map.items()}
    Pm = [P_reverse_map[p] for p in verplichte_posities]

    r_p_k = {}
    matrix_start_index = taken_start_row + 3
    for row in data_rows[matrix_start_index:]:
        positie_naam = row[positie_col]
        if not isinstance(positie_naam, str) or not positie_naam.strip():
            continue
        if str(row[verplicht_col]).strip() != "1":
            continue
        p_index = P_reverse_map.get(positie_naam)
        for i_k, taaknaam in enumerate(K_names):
            col_idx = taken_start_col + i_k
            if col_idx < len(row) and row[col_idx] == 1:
                r_p_k[(p_index, i_k)] = 1

    for base_name in verplichte_posities:
        base_index = P_reverse_map[base_name]
        training_index = P_reverse_map[base_name + " training"]
        for (p_idx, k_idx), val in list(r_p_k.items()):
            if p_idx == base_index:
                r_p_k[(training_index, k_idx)] = val

    return {
        "K": list(K_map.keys()),
        "K_map": K_map,
        "K_reverse_map": K_reverse_map,
        "P": P,
        "P_map": P_map,
        "P_reverse_map": P_reverse_map,
        "Pm": Pm,
        "r_p_k": r_p_k
    }


def load_training_data(filepath: str, sheet_name: str, K_reverse_map: dict, allowed_ids: set[int]) -> dict:
    import os  # zorg dat dit bovenin staat
    import math

    df = pd.read_excel(filepath, sheet_name=sheet_name)
    df.columns = [str(c).strip() for c in df.columns]

    required_columns = {"Fictive EmployeeId", "TaskDescription", "EmployabilityId", "CompletedTrainingInHours", "TrainingLoadInHours"}
    if not required_columns.issubset(df.columns):
        raise ValueError(f"❌ Missende kolommen in sheet '{sheet_name}': {required_columns - set(df.columns)}")

    df = df[df["Fictive EmployeeId"].isin(allowed_ids)]
    df["TaskDescription"] = df["TaskDescription"].astype(str).str.strip().str.upper()

    valid_skills = set(K_reverse_map.keys())
    df = df[df["TaskDescription"].isin(valid_skills)]

    # === STAP 1: blokkeer skills voor (w, k) met ooit employability 0 ===
    blocked_skills = set()
    for _, row in df.iterrows():
        w = int(row["Fictive EmployeeId"])
        k_str = row["TaskDescription"]
        e = int(row.get("EmployabilityId", 0))
        if k_str not in K_reverse_map:
            continue
        k = K_reverse_map[k_str]
        if e == 0:
            blocked_skills.add((w, k))

    # === STAP 2: filter alles wat geblokkeerd is
    df = df[~df.apply(lambda row: (int(row["Fictive EmployeeId"]), K_reverse_map.get(row["TaskDescription"])) in blocked_skills, axis=1)]

    # === STAP 3: opbouw van skills en loads
    missing_training_load = df[df["TrainingLoadInHours"].isna()]
    if not missing_training_load.empty:
        print("⚠️ Volgende skills hebben geen trainingsduur (TrainingLoadInHours):")
        print(missing_training_load["TaskDescription"].value_counts())

    W = sorted(df["Fictive EmployeeId"].dropna().astype(int).unique())
    K_used = sorted({K_reverse_map[sk] for sk in df["TaskDescription"]})

    # OpenTrainingInHours = vereiste load (origineel)
    L_k = df.groupby("TaskDescription")["OpenTrainingInHours"].max().to_dict()
    L_k = {K_reverse_map[k]: v for k, v in L_k.items() if k in K_reverse_map and v > 0}
    L_k_orig = L_k.copy()  # bewaren vóór buffering

    I_wks = {w: {k: {0: 0} for k in K_used} for w in W}
    experts = {w: {k: 0 for k in K_used} for w in W}

    for _, row in df.iterrows():
        try:
            w = int(row["Fictive EmployeeId"])
            k_str = row["TaskDescription"]
            if k_str not in K_reverse_map:
                continue
            k = K_reverse_map[k_str]
            if (w, k) in blocked_skills:
                continue

            e = int(row.get("EmployabilityId", 0))
            ch = float(row.get("CompletedTrainingInHours", 0))
            full_load = L_k.get(k)
            if full_load is None:
                print(f"⚠️ Skill '{k_str}' (id {k}) ontbreekt in L_k → wordt overgeslagen.")
                continue

            if e in [2, 3, 4]:
                I_wks[w][k][0] = full_load
                if e >= 3:
                    experts[w][k] = 1
            elif e == 1:
                I_wks[w][k][0] = ch

        except Exception as ex:
            print(f"⚠️ Fout bij verwerken rij: {row.to_dict()} → {ex}")

    # === STAP 4: trainingbuffer toepassen
    gamma_str = os.environ.get("GAMMA_TRAINING_DURATION", None)
    if gamma_str is not None:
        gamma = float(gamma_str)
        # Schaal eerst alle originele L_k op
        L_k = {k: math.ceil(L_k_orig[k] * gamma) for k in L_k_orig}

        # Pas I_wks aan als training oorspronkelijk volledig was
        for w in I_wks:
            for k in I_wks[w]:
                if k in L_k_orig:
                    current = I_wks[w][k][0]
                    if current >= 0.99 * L_k_orig[k]:
                        I_wks[w][k][0] = current * gamma  # schaal gevolgd volume mee

    return {
        "W": W,
        "K": K_used,
        "I_wks": I_wks,
        "experts": experts,
        "L_k": L_k
    }


def load_training_data_historic(filepath: str, sheet_name: str, K_reverse_map: dict, allowed_ids: set[int], cutoff_date: datetime) -> dict:
    df = pd.read_excel(filepath, sheet_name=sheet_name)
    df.columns = [str(c).strip() for c in df.columns]

    required_columns = {"Fictive EmployeeId", "TaskDescription", "EmployabilityId", "CompletedTrainingInHours", "TrainingLoadInHours", "Datum"}
    if not required_columns.issubset(df.columns):
        raise ValueError(f"❌ Missende kolommen in sheet '{sheet_name}': {required_columns - set(df.columns)}")

    df["Datum"] = pd.to_datetime(df["Datum"], errors="coerce")
    df = df[df["Datum"] <= cutoff_date]
    df = df[df["Fictive EmployeeId"].isin(allowed_ids)]
    df["TaskDescription"] = df["TaskDescription"].astype(str).str.strip().str.upper()

    valid_skills = set(K_reverse_map.keys())
    df = df[df["TaskDescription"].isin(valid_skills)]

    blocked_skills = set()
    for _, row in df.iterrows():
        w = int(row["Fictive EmployeeId"])
        k_str = row["TaskDescription"]
        e = int(row.get("EmployabilityId", 0))
        if k_str not in K_reverse_map:
            continue
        k = K_reverse_map[k_str]
        if e == 0:
            blocked_skills.add((w, k))

    df = df[~df.apply(lambda row: (int(row["Fictive EmployeeId"]), K_reverse_map.get(row["TaskDescription"])) in blocked_skills, axis=1)]

    W = sorted(df["Fictive EmployeeId"].dropna().astype(int).unique())
    K_used = sorted({K_reverse_map[sk] for sk in df["TaskDescription"]})
    L_k_raw = df.groupby("TaskDescription")["TrainingLoadInHours"].first().to_dict()
    L_k = {K_reverse_map[k]: v for k, v in L_k_raw.items() if k in K_reverse_map}

    I_wks = {w: {k: {0: 0} for k in K_used} for w in W}
    experts = {w: {k: 0 for k in K_used} for w in W}

    df = df.sort_values("Datum")  # belangrijk: laatste staat telt!

    for (w, k), group in df.groupby([df["Fictive EmployeeId"], df["TaskDescription"]]):
        w = int(w)
        k_idx = K_reverse_map.get(str(k).strip().upper())
        if (w, k_idx) in blocked_skills:
            continue
        latest = group.iloc[-1]
        e = int(latest.get("EmployabilityId", 0))
        ch = float(latest.get("CompletedTrainingInHours", 0))
        full_load = L_k.get(k_idx)

        if full_load is None:
            continue

        if e in [2, 3, 4]:
            I_wks[w][k_idx][0] = full_load
            if e >= 3:
                experts[w][k_idx] = 1
        elif e == 1:
            I_wks[w][k_idx][0] = ch

    return {
        "W": W,
        "K": K_used,
        "I_wks": I_wks,
        "experts": experts,
        "L_k": L_k
    }


def load_parttime_percentages(filepath: str) -> dict:
    df = pd.read_excel(filepath)
    df = df.dropna(subset=["Fictive EmployeeId", "ProcentParttime"])
    df["Fictive EmployeeId"] = df["Fictive EmployeeId"].astype(int)
    df["ProcentParttime"] = df["ProcentParttime"].astype(int)
    return dict(zip(df["Fictive EmployeeId"], df["ProcentParttime"]))

def load_allowed_ids(filepath: str, team: str) -> set[int]:
    """
    Filtert toegestane werknemer-IDs op basis van 'Gebruik In Telling Kal.' == 1.
    """
    df = pd.read_excel(filepath)
    df.columns = df.columns.str.strip()
    df = df[df["Gebruik In Telling Kal."].isin([1, 1.0, True, "True"])]
    df = df[pd.notna(df["Fictief Stam Nr."])]
    return set(df["Fictief Stam Nr."].astype(int))
def load_verlofsaldo(filepath: str) -> dict[int, float]:
    """
    Laadt totaal verlofsaldo per werknemer-ID.
    """
    df = pd.read_excel(filepath)
    df.columns = df.columns.str.strip()
    df["OPGENOMEN_DAGEN"] = pd.to_numeric(df["OPGENOMEN_DAGEN"], errors="coerce").fillna(0)
    df["NOG_TE_NEMEN_DAGEN"] = pd.to_numeric(df["NOG_TE_NEMEN_DAGEN"], errors="coerce").fillna(0)
    df["TOTAAL_DAGEN"] = df["OPGENOMEN_DAGEN"] + df["NOG_TE_NEMEN_DAGEN"]
    return df.groupby("StamNr Fiktief")["TOTAAL_DAGEN"].sum().to_dict()
def load_simulation_parameters(filepath: str) -> dict:
    """
    Laadt kansverdeling eerste verlofdag en gemiddelde verlofduur.
    """
    df_prob = pd.read_excel(filepath, sheet_name="EersteVerlofPerMaand")
    month_probs = df_prob.set_index("Maand")["Kans op eerste verlofdag (per dag)"].to_dict()

    df_duur = pd.read_excel(filepath, sheet_name="Verlofduur")
    verlofduur = df_duur.iloc[0, 0]

    return {
        "month_probs": month_probs,
        "verlofduur": verlofduur
    }
# Integratie in input_module.py
def load_empirical_samplers(
    afwezigheden_path: str,
    werknemers_path: str,
    use_kde_sick: bool = False,
    use_kde_vacation: bool = True,
    bandwidth_factor_sick: float = 1.0,
    bandwidth_factor_vacation: float = 0.8
) -> tuple:
    """
    Laadt observed sickness en vacation durations en maakt EmpiricalSamplers aan.

    Parameters:
    - afwezigheden_path: pad naar Afwezigheden.xlsx
    - werknemers_path: pad naar Werknemers.xlsx
    - use_kde_sick: gebruik KDE voor sick sampler
    - use_kde_vacation: gebruik KDE voor vacation sampler
    - bandwidth_factor_sick: bandwidth voor KDE sick sampler
    - bandwidth_factor_vacation: bandwidth voor KDE vacation sampler

    Returns:
    - sick_sampler, vacation_sampler
    """
    afw_df = pd.read_excel(afwezigheden_path, sheet_name="data")
    wn_df = pd.read_excel(werknemers_path, sheet_name="data")

    afw_df.columns = afw_df.columns.str.strip()
    wn_df.columns = wn_df.columns.str.strip()
    afw_df = afw_df.rename(columns={"Fikitef StamNr": "Stamnummer", "TIME_ZIEK": "Ziek", "DAG": "Datum"})
    wn_df = wn_df.rename(columns={"StamNr Fiktief": "Stamnummer"})

    afw = afw_df.merge(wn_df[["Stamnummer", "Leeftijd"]], on="Stamnummer", how="left")
    afw["Datum"] = pd.to_datetime(afw["Datum"])

    # Ziektereeksen extraheren
    afw = afw.sort_values(["Stamnummer", "Datum"])
    afw["IsZiek"] = afw["Ziek"] > 0
    afw["groep"] = afw.groupby("Stamnummer")["IsZiek"].diff().ne(0).cumsum()

    ziekte_reeksen = afw[afw["IsZiek"]].groupby(["Stamnummer", "groep"]).size().reset_index(name="ZiekteduurShifts")
    observed_sick_durations = ziekte_reeksen["ZiekteduurShifts"].tolist()

    # Verlofduren extraheren
    afw_df = afw_df[["Stamnummer", "Datum", "TIME_VAK"]].dropna(subset=["Stamnummer", "Datum"])
    afw_df["IsVerlof"] = afw_df["TIME_VAK"] > 0
    afw_df = afw_df.sort_values(["Stamnummer", "Datum"])
    afw_df["groep"] = afw_df.groupby("Stamnummer")["IsVerlof"].diff().ne(0).cumsum()

    verlof_reeksen = afw_df[afw_df["IsVerlof"]].groupby(["Stamnummer", "groep"]).size().reset_index(name="DuurVerlof")
    observed_vacation_durations = verlof_reeksen["DuurVerlof"].tolist()

    # Samplers maken
    sick_sampler = EmpiricalSampler(
        observed_sick_durations,
        use_kde=use_kde_sick,
        bandwidth_factor=bandwidth_factor_sick
    )
    vacation_sampler = EmpiricalSampler(
        observed_vacation_durations,
        use_kde=use_kde_vacation,
        bandwidth_factor=bandwidth_factor_vacation
    )

    return sick_sampler, vacation_sampler


def apply_buffers_to_inputs(inputs: dict, buffer_factors: dict) -> dict:
    """
    Applies buffer factors to the relevant fields inside the inputs dictionary.

    - Training loads (L_k) are rounded UP to the nearest integer
    - Probabilities are rounded normally to 5 decimals
    - Verlofduur is rounded UP
    """

    # 1. Training durations (L_k)
    if "L_k" in inputs and "training_duration" in buffer_factors:
        gamma = buffer_factors["training_duration"]
        inputs["L_k"] = {k: math.ceil(v * gamma) for k, v in inputs["L_k"].items()}

    # 2. Sickness probabilities (ziekte_mapping)
    if "ziekte_mapping" in inputs and "sick_prob" in buffer_factors:
        gamma = buffer_factors["sick_prob"]
        for w in inputs["ziekte_mapping"]:
            p = inputs["ziekte_mapping"][w]["prob_sick"]
            inputs["ziekte_mapping"][w]["prob_sick"] = min(1.0, round(p * gamma, 5))

    # 3. Vacation first day probabilities (sim_params -> month_probs)
    if "sim_params" in inputs and "vacation_prob" in buffer_factors:
        gamma = buffer_factors["vacation_prob"]
        inputs["sim_params"]["month_probs"] = {m: min(1.0, round(p * gamma, 5)) for m, p in
                                               inputs["sim_params"]["month_probs"].items()}

    # 4. Mutation probability (sim_params -> mutation_prob)
    if "sim_params" in inputs and "mutation_prob" in buffer_factors:
        gamma = buffer_factors["mutation_prob"]
        if "mutation_prob" in inputs["sim_params"]:
            p = inputs["sim_params"]["mutation_prob"]
            inputs["sim_params"]["mutation_prob"] = min(1.0, round(p * gamma, 5))

    # 5. Vacation mean duration (sim_params -> verlofduur)
    if "sim_params" in inputs and "vacation_duration" in buffer_factors:
        gamma = buffer_factors["vacation_duration"]
        inputs["sim_params"]["verlofduur"] = math.ceil(inputs["sim_params"]["verlofduur"] * gamma)

    return inputs


def load_all_inputs(start_date,
                    end_date,
                    team_name="Ploeg D",
                    GAMMA_SICK_DURATION = 1.0,
                    GAMMA_VACATION_DURATION = 1.0,
                    GAMMA_TRAINING_DURATION = 1.0,
                    GAMMA_SICK_PROB = 1.0,
                    GAMMA_VACATION_PROB = 1.0,
                    GAMMA_MUTATION_PROB = 1.0) -> dict:
    # === Buffer dictionary opbouwen uit argumenten
    buffer_factors = {
        "training_duration": GAMMA_TRAINING_DURATION,
        "sick_duration": GAMMA_SICK_DURATION,
        "vacation_duration": GAMMA_VACATION_DURATION,
        "sick_prob": GAMMA_SICK_PROB,
        "vacation_prob": GAMMA_VACATION_PROB,
        "mutation_prob": GAMMA_MUTATION_PROB
    }

    comp = load_competence_structure(
        filepath=r"C:\Users\olafs\OneDrive\Bureaublad\master's dissertation\variation and competences\AM_Data\SPL_SDG_competentiematrix_VOLLEDIG_stamnummers.xlsx",
        sheet_name = f"SDG 1 {team_name}"
    )
    allowed_ids = load_allowed_ids(
        filepath=rf"C:\Users\olafs\OneDrive\Bureaublad\master's dissertation\variation and competences\AM_Data\RE_ extra data taken\SDG 1 - {team_name} - wn inzetbaarheden.xlsx",
        team= f"{team_name}"
    )

    training = load_training_data(
        filepath=r"C:\Users\olafs\OneDrive\Bureaublad\master's dissertation\variation and competences\AM_Data\Trainings.xlsx",
        sheet_name=f"{team_name.upper()} SDG1",
        K_reverse_map=comp["K_reverse_map"],
        allowed_ids=allowed_ids
    )
    verlof_targets = load_verlofsaldo(r"C:\Users\olafs\OneDrive\Bureaublad\master's dissertation\variation and competences\AM_Data\VerlofSaldo.xlsx")
    sim_params = load_simulation_parameters("SimulatieResultaten.xlsx")

    S = list(range(int((end_date - start_date).days * 0.75)))
    comp["K_p"] = create_K_p_from_r_p_k(comp["r_p_k"])

    # ❗ Eerst buffers toepassen vóór simulatie
    raw_inputs = {
        **comp,
        **training,
        "allowed_ids": allowed_ids,
        "verlofsaldo": verlof_targets,
        "sim_params": sim_params
    }
    raw_inputs = apply_buffers_to_inputs(raw_inputs, buffer_factors)

    S = list(range(int((end_date - start_date).days * 0.75)))

    sick_sampler, vacation_sampler = load_empirical_samplers(
        afwezigheden_path=r"C:\Users\olafs\OneDrive\Bureaublad\master's dissertation\variation and competences\AM_Data\Afwezigheden.xlsx",
        werknemers_path=r"C:\Users\olafs\OneDrive\Bureaublad\master's dissertation\variation and competences\AM_Data\Werknemers.xlsx",
        use_kde_sick=False,
        use_kde_vacation=True,
        bandwidth_factor_vacation=0.8
    )

    vacation_plan, vacation_matrix, shift_leave_map = simulate_vacation_days(
        W=sorted(allowed_ids),
        start_date=start_date,
        end_date=end_date,
        leave_targets=raw_inputs["verlofsaldo"],
        month_probs=raw_inputs["sim_params"]["month_probs"],
        vacation_sampler=vacation_sampler,
        mean_leave_duration=raw_inputs["sim_params"]["verlofduur"],
        min_available=8
    )


    f_w_s = build_f_w_s_from_vacation_matrix(vacation_matrix, S)

    alfa_p = load_alfa_p_weights_from_excel("alfa_weights.xlsx")
    if alfa_p is None:
        alfa_p = {p: 1 for p in comp["P"]}

    return {
        **raw_inputs,
        "vacation_plan": vacation_plan,
        "vacation_matrix": vacation_matrix,
        "f_w_s": f_w_s,
        "alfa_p": alfa_p,
        "S": S,
        "start_date": start_date,
        "end_date": end_date,
        "sick_sampler": sick_sampler,
        "vacation_sampler": vacation_sampler
    }


def build_f_w_s_from_vacation_matrix(vacation_matrix: pd.DataFrame, shifts: list[int]) -> dict[int, dict[int, int]]:
    """
    Bouwt een f_w_s matrix: 1 = werknemer is afwezig (verlof), 0 = beschikbaar,
    gebaseerd op een vakantie-matrix.

    Parameters:
    - vacation_matrix: DataFrame (index = werknemers, kolommen = shifts)
    - shifts: lijst van shifts (bv. S = list(range(N)))

    Returns:
    - dict f_w_s[w][s] = 1 of 0
    """
    f_w_s = {}

    for w in vacation_matrix.index:
        f_w_s[w] = {}
        for s in shifts:
            try:
                val = vacation_matrix.at[w, s]
            except KeyError:
                # Shift bestaat niet (bijvoorbeeld matrix kleiner dan S)
                f_w_s[w][s] = 0
                continue

            if isinstance(val, str) and val.strip().lower() == "x":
                f_w_s[w][s] = 1  # Niet beschikbaar
            else:
                f_w_s[w][s] = 0  # Beschikbaar
    export_path = os.path.join("output", "vacation_matrix_fws.xlsx")
    export_f_w_s_to_excel(f_w_s,export_path)
    return f_w_s
def export_f_w_s_to_excel(f_w_s: dict[int, dict[int, int]], file_path: str):
    import pandas as pd
    df = pd.DataFrame.from_dict(f_w_s, orient="index")
    df.index.name = "Worker"
    df.to_excel(file_path)
    print(f"✅ f_w_s geëxporteerd naar: {file_path}")

def build_sickness_mapping(
    employee_file: str,
    sim_file: str,
    employee_sheet: str = "data",
    sickness_sheet: str = "Ziekteanalyse",
    default_age: int = 35
) -> dict[int, dict]:


    # Hulpfunctie om leeftijdscategorie te bepalen
    def get_leeftijdsklasse(age: int) -> str:
        if age < 30:
            return "<30"
        elif age < 40:
            return "30-40"
        elif age < 50:
            return "40-50"
        else:
            return "50+"

    # Laad data
    df_emp = pd.read_excel(employee_file, sheet_name=employee_sheet)
    df_sim = pd.read_excel(sim_file, sheet_name=sickness_sheet)

    # Zorg dat kolomnamen juist zijn
    df_emp = df_emp.dropna(subset=["StamNr Fiktief", "Leeftijd"])
    df_emp["Stamnr"] = df_emp["StamNr Fiktief"].astype(int)
    df_emp["Leeftijd"] = df_emp["Leeftijd"].astype(int)
    df_emp["Leeftijdsklasse"] = df_emp["Leeftijd"].apply(get_leeftijdsklasse)

    # Zet kolomnamen ziekteanalyse gelijk
    df_sim.columns = df_sim.columns.str.strip()
    if "Leeftijdsklasse" not in df_sim.columns:
        df_sim.rename(columns={df_sim.columns[0]: "Leeftijdsklasse"}, inplace=True)

    df_sim.set_index("Leeftijdsklasse", inplace=True)

    # Fallbackwaarden voor default leeftijd
    fallback_klasse = get_leeftijdsklasse(default_age)
    fallback = df_sim.loc[fallback_klasse].to_dict()

    # Mapping bouwen
    ziekte_map = {}
    for _, row in df_emp.iterrows():
        stamnr = row["Stamnr"]
        klasse = row["Leeftijdsklasse"]

        if klasse in df_sim.index:
            data = df_sim.loc[klasse].to_dict()
        else:
            data = fallback

        ziekte_map[stamnr] = {
            "prob_sick": data.get("Ziektekans", 0.01),
            "mean_duration": data.get("Gemiddelde duur", 5.0),
            "std_duration": data.get("Standaardafwijking duur", 8.0)
        }

    return ziekte_map


def load_mutation_probability(sim_file="SimulatieResultaten.xlsx", sheet="Mutatieanalyse"):


    df = pd.read_excel(sim_file, sheet_name=sheet)
    if "Kans per shift per persoon" in df.columns:
        prob = df.iloc[1, 1]  # rij 2, kolom 2 = B2
        return float(prob)
    else:
        raise ValueError(f"Kolom 'Mutatieanalyse' niet gevonden in sheet {sheet}")

def save_alfa_p_weights_to_excel(alfa_p: dict[int, float], export_path: str):
    """
    Slaat de gewichten per positie (`alfa_p`) op in een Excelbestand, om later opnieuw in te laden.

    Parameters:
    - alfa_p: dictionary met positie-ID → gewicht
    - export_path: pad naar Excelbestand (bijv. 'alfa_weights.xlsx')
    """
    df = pd.DataFrame([
        {"PositionID": p, "Weight": w}
        for p, w in alfa_p.items()
    ])
    df.to_excel(export_path, index=False)

def load_alfa_p_weights_from_excel(filepath: str) -> dict[int, float]:
    """
    Laadt een Excelbestand met gewichten per positie.

    Verwacht kolommen: 'PositionID' en 'Weight'.

    Returns:
    - Dictionary: {positie-ID: gewicht}
    """
    if not os.path.exists(filepath):
        return None  # fallback laten beslissen in hoofdscript

    df = pd.read_excel(filepath)
    return dict(zip(df["PositionID"], df["Weight"]))

def create_K_p_from_r_p_k(r_p_k: dict) -> dict:
    """Create K_p mapping (position → list of required skills) from r_p_k"""
    K_p = defaultdict(list)
    for (p, k), val in r_p_k.items():
        if val == 1:  # Only include required skills
            K_p[p].append(k)
    return dict(K_p)

